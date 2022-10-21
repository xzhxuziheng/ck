"""
coder: xuziheng
date: 2022/10/13 11:25
"""
from openpyxl import load_workbook


class ExcelData:
    def __init__(self, excel_path, excel_sheet):
        self.excel_path = excel_path
        self.excel_sheet = excel_sheet

    def read_excel(self):
        wb = load_workbook(self.excel_path)
        ws = wb[self.excel_sheet]
        max_row = ws.max_row
        max_col = ws.max_column
        for row in range(2, max_row+1):
            data = {}
            for col in range(1, max_col+1):
                data[ws.cell(row=1, column=col).value] = ws.cell(row=row, column=col).value
            yield data


if __name__ == '__main__':
    e = ExcelData('./告警数据依赖.xlsx', 'Sheet1')
    print([i for i in e.read_excel()])
