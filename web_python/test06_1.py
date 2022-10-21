"""
coder: xuziheng
date: 2022/9/6 16:07
"""


class Filed:
    def __get__(self, instance, owner):
        print('---调用获取属性---')
        return self.value

    def __set__(self, instance, value):
        print('---调用设置属性---')
        if value == 'attr':
            print('打印：', value)
            self.value = value
        else:
            self.value = 1900

    def __delete__(self, instance):
        pass
    

class Model:
    attr = Filed()
    name = 'test'


# m = Model()
# m.age = 20
# m.attr = 1000
# print(m.name)
# print(m.age)
# print(m.attr)
from openpyxl import load_workbook


class Excel:

    def read_excel(self):
        wb = load_workbook('test02.xlsx')
        # wb.create_sheet('test')
        # wb.save('test02.xlsx')
        ws = wb['Sheet1']
        max_row = ws.max_row
        max_col = ws.max_column

        for row in range(2, max_row+1):
            data = []
            data_dict = {}
            for col in range(1, max_col+1):
                data_dict[ws.cell(row=1, column=col).value] = ws.cell(row=row, column=col).value
            data.append(data_dict)
        # ws.cell(row=12, column=1).value = 'write test test1'
        # wb.save('test02.xlsx')
        # print(ws.cell(row=12, column=1).value)
            yield data_dict

# t = Excel()
# print(t.read_excel())
# print(list(t.read_excel()))





















# class Exec:
#
#     def __init__(self, excel_path, excel_sheet):
#         self.excel_path = excel_path
#         self.excel_sheet = excel_sheet
#
#     def read_excel(self):
#         wb = load_workbook(self.excel_path)
#         ws = wb[self.excel_sheet]
#         max_row = ws.max_row
#         max_col = ws.max_column
#         for row in range(2, max_row+1):
#             data_excel = {}
#             for col in range(1, max_col+1):
#                 data_excel[ws.cell(row=1, column=col).value] = ws.cell(row=row, column=col).value
#             yield data_excel
#
#
# t = Exec('test02.xlsx', 'Sheet1')
# data = list(t.read_excel())
# print(data[0])












