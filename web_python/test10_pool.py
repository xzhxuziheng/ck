"""
coder: xuziheng
date: 2022/9/19 11:34
"""

from openpyxl import load_workbook
from multiprocessing import Queue, Pool, Manager
import time
import requests


def read_excel(excel_path, sheet_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    row_max = ws.max_row
    col_max = ws.max_column
    for row in range(2, row_max+1):
        data = {}
        for col in range(1, col_max+1):
            data[ws.cell(row=1, column=col).value] = ws.cell(row=row, column=col).value
        yield data


# 进程池中的队列（给进程池中的各个进程之间使用）
i = 0


def work(q):
    global i
    while q.qsize() > 0:
        url = q.get()
        header = {
            'content-type': 'application/json',
            'area_id': '7',
            'source_type': '3',
            'token': 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJzb3VyY2VfdHlwZSI6Mywic291cmNlX2lkIjoxMjE4fQ.APKg0vDbJvVqJEiwFBmB73dV_eEOfZX7p87z9HXoIp4'
        }
        requests.get(url=url, headers=header)

        i += 1
    print('该进程运行了{}次'.format(i))


if __name__ == '__main__':
    q3 = Manager().Queue()
    for i in range(10):
        q3.put('https://test.bike.ledear.cn/api/user/refund-apply/list?page=1&pageSize=10')
    pool = Pool(3)
    st = time.time()
    for i in range(3):
        pool.apply_async(work, args=(q3,))
    pool.close()
    pool.join()
    et = time.time()
    print('耗时：{}'.format(et-st))


