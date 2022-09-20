"""
coder: xuziheng
date: 2022/8/25 15:03
"""
from openpyxl import load_workbook


# def read_excel(file_path, sheet_name):
#     """
#     读取excel的方法
#     :param file_path: 文件路径
#     :param sheet_name: 表名
#     :return: 读取的数据
#     """
#     workbook = load_workbook(file_path)
#     sheet = workbook[sheet_name]
#     row_max = sheet.max_row
#     column_max = sheet.max_column
#     for row in range(2, row_max+1):
#         # 每行数据
#         row_dict = {}
#         for col in range(1, column_max+1):
#             row_dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
#         yield row_dict
# print(list(read_excel('test02.xlsx', 'Sheet1')))


def gen_excel(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    row_max = ws.max_row + 1
    col_max = ws.max_column + 1
    # 前面一个for j循环是内层，后面一个for i循环是外层
    test_data = ({ws.cell(1, j).value: ws.cell(i, j).value for j in range(1, col_max)} for i in range(2, row_max))
    return test_data


g = gen_excel('test02.xlsx', 'Sheet1')
print(g)
print(list(g))
for i in g:
    print(i)

